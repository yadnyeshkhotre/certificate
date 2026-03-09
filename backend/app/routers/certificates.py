from datetime import date, datetime
from io import BytesIO
import base64
import json
import re
import zipfile
from typing import Any, Literal

from fastapi import APIRouter, File, Form, HTTPException, UploadFile, status
from fastapi.responses import Response
from openpyxl import load_workbook

from ..schemas import (
    BulkFailure,
    BulkCertificateDownloadRequest,
    BulkGenerationResponse,
    CertificatePayload,
    CertificateRecord,
    GenerateCertificateRequest,
)
from ..services.certificate_service import create_certificate_record
from ..services.certificate_render_service import render_certificate_image
from ..store import append_certificates, get_certificate, get_template, load_certificates

router = APIRouter(prefix='/certificates', tags=['certificates'])


def _safe_name_component(value: str) -> str:
    collapsed = re.sub(r'\s+', '_', value.strip())
    cleaned = re.sub(r'[^A-Za-z0-9._-]', '', collapsed)
    return cleaned or 'recipient'


@router.get('', response_model=list[CertificateRecord])
def list_certificates() -> list[CertificateRecord]:
    return load_certificates()


@router.get('/{certificate_id}', response_model=CertificateRecord)
def read_certificate(certificate_id: str) -> CertificateRecord:
    certificate = get_certificate(certificate_id)
    if certificate is None:
        raise HTTPException(status_code=404, detail='Certificate not found.')
    return certificate


@router.get('/{certificate_id}/download')
def download_certificate(
    certificate_id: str,
    format: Literal['png', 'jpg'] = 'png',
) -> Response:
    certificate = get_certificate(certificate_id)
    if certificate is None:
        raise HTTPException(status_code=404, detail='Certificate not found.')

    image_bytes = render_certificate_image(certificate, image_format=format)
    extension = 'jpg' if format == 'jpg' else 'png'
    media_type = 'image/jpeg' if extension == 'jpg' else 'image/png'
    filename = f'certificate-{certificate.certificate_id}.{extension}'

    return Response(
        content=image_bytes,
        media_type=media_type,
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


@router.post('/bulk/download')
def download_bulk_certificates(request: BulkCertificateDownloadRequest) -> Response:
    certificates = {record.certificate_id: record for record in load_certificates()}
    unique_ids = list(dict.fromkeys(request.certificate_ids))

    missing_ids = [certificate_id for certificate_id in unique_ids if certificate_id not in certificates]
    if missing_ids:
        raise HTTPException(
            status_code=404,
            detail=f'Certificate not found: {missing_ids[0]}',
        )

    extension = 'jpg' if request.format == 'jpg' else 'png'
    zip_buffer = BytesIO()
    manifest_rows: list[dict[str, str]] = []

    with zipfile.ZipFile(zip_buffer, mode='w', compression=zipfile.ZIP_DEFLATED) as archive:
        for index, certificate_id in enumerate(unique_ids, start=1):
            record = certificates[certificate_id]
            image_bytes = render_certificate_image(record, image_format=request.format)
            recipient = _safe_name_component(record.payload.recipient_name)
            file_name = f'{index:03d}_{recipient}_{certificate_id[:8]}.{extension}'
            archive.writestr(file_name, image_bytes)

            manifest_rows.append(
                {
                    'certificate_id': record.certificate_id,
                    'recipient_name': record.payload.recipient_name,
                    'course_name': record.payload.course_name,
                    'issue_date': record.payload.issue_date.isoformat(),
                    'file_name': file_name,
                }
            )

        archive.writestr(
            'manifest.json',
            json.dumps({'certificates': manifest_rows}, indent=2),
        )

    zip_filename = f'bulk-certificates-{datetime.now().strftime("%Y%m%d-%H%M%S")}.zip'
    return Response(
        content=zip_buffer.getvalue(),
        media_type='application/zip',
        headers={
            'Content-Disposition': f'attachment; filename="{zip_filename}"',
            'X-Generated-Count': str(len(unique_ids)),
        },
    )


@router.post('/generate', response_model=CertificateRecord, status_code=status.HTTP_201_CREATED)
def generate_certificate(request: GenerateCertificateRequest) -> CertificateRecord:
    if get_template(request.template_id) is None:
        raise HTTPException(status_code=404, detail='Template not found.')

    certificate = create_certificate_record(request.template_id, request.payload)
    append_certificates([certificate])
    return certificate


def _parse_issue_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        candidate = value.strip()
        if not candidate:
            raise ValueError('issue_date is required.')
        try:
            return date.fromisoformat(candidate)
        except ValueError as exc:
            raise ValueError('issue_date must be ISO format YYYY-MM-DD.') from exc
    raise ValueError('issue_date must be an Excel date or YYYY-MM-DD string.')


def _normalize_row(headers: list[str], row: tuple[Any, ...]) -> dict[str, Any]:
    payload: dict[str, Any] = {}
    for index, header in enumerate(headers):
        if not header:
            continue
        payload[header] = row[index] if index < len(row) else None
    return payload


def _build_payload_from_row(
    row_payload: dict[str, Any],
    default_issuer_name: str | None = None,
    default_issuer_signature_data_url: str | None = None,
) -> CertificatePayload:
    recipient_name = str(row_payload.get('recipient_name', '') or '').strip()
    course_name = str(row_payload.get('course_name', '') or '').strip()

    if not recipient_name:
        raise ValueError('recipient_name is required.')
    if not course_name:
        raise ValueError('course_name is required.')

    issuer_value = row_payload.get('issuer_name')
    issuer_name = str(issuer_value).strip() if issuer_value is not None else default_issuer_name
    if issuer_name == '':
        issuer_name = None

    issuer_signature_value = row_payload.get('issuer_signature_data_url')
    issuer_signature_data_url = (
        str(issuer_signature_value).strip()
        if issuer_signature_value is not None
        else default_issuer_signature_data_url
    )
    if issuer_signature_data_url == '':
        issuer_signature_data_url = None

    issue_date = _parse_issue_date(row_payload.get('issue_date'))

    metadata: dict[str, Any] = {}
    core_fields = {
        'recipient_name',
        'course_name',
        'issue_date',
        'issuer_name',
        'issuer_signature_data_url',
    }
    for key, value in row_payload.items():
        if key in core_fields or value is None:
            continue
        metadata[key] = value

    return CertificatePayload(
        recipient_name=recipient_name,
        course_name=course_name,
        issue_date=issue_date,
        issuer_name=issuer_name,
        issuer_signature_data_url=issuer_signature_data_url,
        metadata=metadata,
    )


async def _read_signature_upload(upload: UploadFile | None) -> str | None:
    if upload is None:
        return None

    filename = (upload.filename or '').lower()
    content_type = (upload.content_type or '').lower()
    if not filename.endswith('.png') and content_type != 'image/png':
        raise HTTPException(status_code=400, detail='Issuer signature must be a PNG file.')

    content = await upload.read()
    if not content:
        raise HTTPException(status_code=400, detail='Issuer signature file is empty.')
    if len(content) > 2 * 1024 * 1024:
        raise HTTPException(status_code=400, detail='Issuer signature PNG must be smaller than 2MB.')
    if content[:8] != b'\x89PNG\r\n\x1a\n':
        raise HTTPException(status_code=400, detail='Issuer signature file is not a valid PNG.')

    encoded = base64.b64encode(content).decode('ascii')
    return f'data:image/png;base64,{encoded}'


@router.post('/bulk', response_model=BulkGenerationResponse, status_code=status.HTTP_201_CREATED)
async def generate_bulk_certificates(
    template_id: str = Form(...),
    issuer_name: str | None = Form(None),
    issuer_signature_file: UploadFile | None = File(None),
    file: UploadFile = File(...),
) -> BulkGenerationResponse:
    if get_template(template_id) is None:
        raise HTTPException(status_code=404, detail='Template not found.')

    filename = file.filename or ''
    if not filename.lower().endswith('.xlsx'):
        raise HTTPException(status_code=400, detail='Only .xlsx files are supported.')

    default_issuer_name = issuer_name.strip() if issuer_name else None
    if default_issuer_name == '':
        default_issuer_name = None
    default_signature_data_url = await _read_signature_upload(issuer_signature_file)

    workbook_bytes = await file.read()
    try:
        workbook = load_workbook(filename=BytesIO(workbook_bytes), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail='Unable to parse Excel file.') from exc

    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail='Excel file is empty.')

    headers = [str(value).strip().lower() if value is not None else '' for value in rows[0]]
    required_headers = {'recipient_name', 'course_name', 'issue_date'}
    missing_headers = [header for header in required_headers if header not in headers]
    if missing_headers:
        missing = ', '.join(sorted(missing_headers))
        raise HTTPException(status_code=400, detail=f'Missing required headers: {missing}')

    generated: list[CertificateRecord] = []
    failed_rows: list[BulkFailure] = []

    for row_number, row in enumerate(rows[1:], start=2):
        if all(cell is None for cell in row):
            continue
        try:
            row_payload = _normalize_row(headers, row)
            payload = _build_payload_from_row(
                row_payload,
                default_issuer_name=default_issuer_name,
                default_issuer_signature_data_url=default_signature_data_url,
            )
            generated.append(create_certificate_record(template_id, payload))
        except ValueError as exc:
            failed_rows.append(BulkFailure(row_number=row_number, reason=str(exc)))

    if generated:
        append_certificates(generated)

    return BulkGenerationResponse(
        generated_count=len(generated),
        certificates=generated,
        failed_rows=failed_rows,
    )
